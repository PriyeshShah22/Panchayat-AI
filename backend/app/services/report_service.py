"""Report generation: PDF (reportlab) and Excel (openpyxl)."""
from __future__ import annotations

import io
import os
from datetime import datetime
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from app.core.config import settings
from app.models.bill import Bill


def generate_bill_pdf(bill: Bill, flat_display: str, resident_display: str) -> bytes:
    """Generate a PDF bill as raw bytes."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("title", parent=styles["Heading1"], alignment=1, spaceAfter=10)
    h2 = ParagraphStyle("h2", parent=styles["Heading3"], spaceBefore=10, spaceAfter=6)

    story = []
    story.append(Paragraph(f"{settings.COMPANY_NAME} — Maintenance Bill", title))
    story.append(Paragraph(f"Bill #: <b>{bill.bill_number}</b>", styles["Normal"]))
    story.append(Paragraph(f"Title: {bill.title}", styles["Normal"]))
    story.append(Spacer(1, 6))
    story.append(Paragraph("Billed To", h2))
    story.append(Paragraph(f"{resident_display}<br/>Flat: {flat_display}", styles["Normal"]))
    story.append(Paragraph("Period / Details", h2))
    if bill.description:
        story.append(Paragraph(bill.description, styles["Normal"]))

    rows = [
        ["Description", "Amount"],
        ["Base maintenance", f"₹ {bill.amount:,.2f}"],
        ["Late fee", f"₹ {bill.late_fee:,.2f}"],
        ["Total", f"₹ {bill.total_amount:,.2f}"],
        ["Paid", f"₹ {bill.paid_amount:,.2f}"],
        ["Outstanding", f"₹ {bill.outstanding:,.2f}"],
    ]
    t = Table(rows, colWidths=[110 * mm, 50 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F62FE")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"Issue date: {bill.issue_date.isoformat()} &nbsp;&nbsp; Due: {bill.due_date.isoformat()}",
                           styles["Normal"]))
    story.append(Paragraph(f"Status: <b>{bill.status.value.upper()}</b>", styles["Normal"]))

    doc.build(story)
    return buf.getvalue()


def _save_bytes(folder: str, name: str, data: bytes) -> str:
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, name)
    with open(path, "wb") as f:
        f.write(data)
    return path


def save_bill_pdf(bill: Bill, flat_display: str, resident_display: str) -> str:
    return _save_bytes(settings.UPLOAD_DIR, f"{bill.bill_number}.pdf",
                       generate_bill_pdf(bill, flat_display, resident_display))


def generate_excel_report(rows: Iterable[dict], columns: list, sheet_name: str = "Report") -> bytes:
    """Build an .xlsx from an iterable of dict rows + column order."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_fill = PatternFill("solid", fgColor="0F62FE")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, col in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col, ""))

    for col_idx, col in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            12, min(40, len(str(col)) + 4)
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_excel_report(rows, columns, name: str, sheet_name: str = "Report") -> str:
    return _save_bytes(settings.UPLOAD_DIR, name,
                       generate_excel_report(rows, columns, sheet_name))
